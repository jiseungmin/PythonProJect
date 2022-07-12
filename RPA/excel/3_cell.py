from openpyxl import Workbook
from random import *
wb = Workbook()
ws = wb.active
ws.title = "god"

ws["A1"] = 1
ws["A2"] = 3
ws["B3"] = 9

print(ws["A1"])  # A1 셀의 정보를 출력
print(ws["A1"].value)  # A1 셀의 값을 출력 or 셀의 값이 없을때 none 출력

# row = 1, 2, 3, 4
#column = a(), b(), c()
print(ws.cell(column=1, row=2).value)  # ws["A2"].value
c = ws.cell(column=3, row=2, value=10)  # ws["C2"].value = 10
print(c.value)

index = 1
# 반복문을 이용하여 숫자 넣기
for x in range(1, 11):  # 1 ~ 10 개의 row의 값
    for y in range(1, 11):
        #ws.cell(column=y, row=x, value=randint(1, 100))
        ws.cell(column=y, row=x, value=index)
        index += 1

wb.save("simple.xlsx")
