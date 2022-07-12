from openpyxl import load_workbook  # 파일 불러오기
wb = load_workbook("simple.xlsx")  # simple.xlsx 파일에서 wb을 불러옴
ws = wb.active  # 활성화된 Sheet

# cell 데이터 불러오기
for x in range(1, 11):  # 1 ~ 10 개의 row의 값
    for y in range(1, 11):
        #ws.cell(column=y, row=x, value=randint(1, 100))
        print(ws.cell(column=y, row=x).value, end=" ")
    print()


# cell 갯수를 모를때
for x in range(1, ws.max_row + 1):
    for y in range(1, ws.max_column + 1):
        print(ws.cell(row=x, column=y).value, end=" ")
    print()
