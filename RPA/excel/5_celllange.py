from openpyxl import Workbook
from random import *
from openpyxl.utils.cell import coordinate_from_string

wb = Workbook()
ws = wb.active

# 데이터 한 줄씩 넣기
ws.append(["번호", "영어", "수학"])
for i in range(1, 11):
    ws.append([i, randint(1, 100), randint(1, 100)])

col_b = ws["B"]  # 영어 칼럼만 가져오기
print(col_b)
for cell in col_b:
    print(cell.value, end=" ")

col_range = ws["B:C"]  # 영어 수학 칼럼 가져오기
for cols in col_range:
    for cell in cols:
        print(cell.value)

row_title = ws[1]  # 1번쨰 row만 가져오기
for sell in row_title:
    print(sell.value)

# row_range = ws[2:6]
# for row in row_range:
#     for cell in row:
#         print(cell.value, end=" ")
#     print()


row_range = ws[2:ws.max_row]  # 2번재 줄 부터 마지막 줄까지
for row in row_range:
    for cell in row:
        print(cell.coordinate, end=" ")
        xy = coordinate_from_string(cell.coordinate)
        print(xy, end=" ")
        print(xy[0], end=" ")  # A
        print(xy[1], end=" ")  # 1
    print()

wb.save("simple.xlsx")
