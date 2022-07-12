from openpyxl import Workbook
wb = Workbook()
# wb.activate
ws = wb.create_sheet()  # 새로운 이름 sheet 기본 이름 설정
ws.title = "Mysheet"  # 시트 이름 변경
ws.sheet_properties.tabColor = "ff66ff"  # 시트 색깔 입히기

ws1 = wb.create_sheet("YourSheet")  # 주어진 이름으로 sheet 생성
ws2 = wb.create_sheet("NewSheet", 2)   # 2번째 index에 sheet 생성

new_ws = wb["NewSheet"]  # Dict 형태로 SHeet 접근
print(wb.sheetnames)  # 모든 이름 확인

# sheet 복사
new_ws["A1"] = "test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied sheet"

wb.save('test.xlsx')
wb.close()
